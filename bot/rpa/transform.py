from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


class TransformError(Exception):
    pass


@dataclass(frozen=True)
class ReportRecord:
    cedula: str
    monto: str
    plazo: str
    fecha: str


@dataclass(frozen=True)
class TransformResult:
    linix_file: Path
    records: list[ReportRecord]


REQUIRED_COLUMNS = {
    "IDENTIFICACION": "cedula",
    "MONTO": "monto",
    "PLAZO": "plazo",
    "FECHASOLICITUD": "fecha",
}


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().upper().replace(" ", "")


def _normalize_digits(value: object, field_name: str) -> str:
    if value is None:
        raise TransformError(f"Campo '{field_name}' vacio")
    if isinstance(value, (int, float)):
        return str(int(round(value)))
    text = str(value).strip()
    digits = re.sub(r"\D", "", text)
    return digits if digits else text


def _format_date(value: object) -> str:
    if value is None:
        raise TransformError("Fecha vacia")
    if isinstance(value, datetime):
        return value.strftime("%d%m%Y")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).strftime("%d%m%Y")
    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value)
            return dt.strftime("%d%m%Y")
        except Exception:
            pass
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d%m%Y")
        except ValueError:
            continue
    raise TransformError(f"Formato de fecha no soportado: {value}")


def _find_header_row(rows: Iterable[tuple]) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows, start=1):
        headers = [_normalize_header(cell) for cell in row]
        header_map = {name: i for i, name in enumerate(headers) if name}
        if all(col in header_map for col in REQUIRED_COLUMNS):
            return idx, header_map
    raise TransformError("No se encontraron columnas requeridas en el XLSX.")


def _read_records(input_path: Path) -> list[ReportRecord]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active

    header_row_idx, header_map = _find_header_row(ws.iter_rows(max_row=20, values_only=True))

    records: list[ReportRecord] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        cedula = _normalize_digits(row[header_map["IDENTIFICACION"]], "Identificacion")
        monto = _normalize_digits(row[header_map["MONTO"]], "Monto")
        plazo = _normalize_digits(row[header_map["PLAZO"]], "Plazo")
        fecha = _format_date(row[header_map["FECHASOLICITUD"]])
        records.append(
            ReportRecord(
                cedula=cedula,
                monto=monto,
                plazo=plazo,
                fecha=fecha,
            )
        )

    if not records:
        raise TransformError("El XLSX no tiene filas de datos.")
    return records


def _write_linix_file(
    records: list[ReportRecord],
    output_dir: Path,
    output_encoding: str,
    periodicidad: str,
) -> Path:
    output_path = output_dir / "cargue linix produccion.csv"
    with output_path.open("w", encoding=output_encoding, newline="\n") as output_file:
        for record in records:
            fields = [
                record.cedula,
                record.monto,
                record.plazo,
                record.fecha,
                "",
                "",
                periodicidad,
                "",
            ]
            output_file.write("|".join(fields) + "\n")
    return output_path


def transform_file(
    input_path: Path,
    output_dir: Path,
    output_encoding: str,
    periodicidad: str,
) -> TransformResult:
    logger = logging.getLogger("rpa")
    records = _read_records(input_path)
    linix_path = _write_linix_file(records, output_dir, output_encoding, periodicidad)
    logger.info("Transformed file saved: %s", linix_path)
    return TransformResult(linix_file=linix_path, records=records)
